"""
interfaz.py — GUI del Analizador Proteome Profiler ARY009
Semillero Biomédica — Detección de Cáncer, ITM
"""
import os
import threading
import traceback

import customtkinter as ctk
from tkinter import filedialog, messagebox

import matplotlib
matplotlib.use("TkAgg")
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
import numpy as np

from motor_apoptosis import (
    AnalizadorGrillaARY009,
    calcular_fold_change,
    promediar_cuantificaciones,
    calcular_fold_change_con_error,
)
from mapa_array import (PRO_APOPTOTICAS, ANTI_APOPTOTICAS, PROTEINAS_UNICAS,
                         PARES_DUPLICADOS)

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")


# ── Descripciones simples (para personas no especialistas) ──────────
PROTEINAS_CLAVE = {
    "Bax": "Activa la muerte celular abriendo huecos en la mitocondria. Que suba es bueno contra el cáncer.",
    "Bcl-2": "Protege a la célula cancerosa de morir. Que baje es bueno contra el cáncer.",
    "Bcl-x": "Protege a la célula cancerosa de morir. Niveles altos = la célula resiste a la quimioterapia.",
    "Cleaved Caspase-3": "Es la 'tijera' que ejecuta la muerte celular. Si aparece, la célula está muriendo.",
    "Pro-Caspase-3": "Es la versión inactiva de la 'tijera'. Cuando baja, significa que se está activando.",
    "Cytochrome c": "Se libera de la mitocondria para iniciar la muerte celular. Que suba es buena señal.",
    "SMAC/Diablo": "Desactiva los protectores de la célula. Que suba es bueno contra el cáncer.",
    "XIAP": "Es un protector que bloquea la muerte celular. Que baje es bueno contra el cáncer.",
    "Survivin": "Hace que las células cancerosas sobrevivan. Que baje es bueno contra el cáncer.",
    "FADD": "Es una pieza clave para que la muerte celular se active desde fuera de la célula.",
    "TRAIL R1/DR4": "Receptor que recibe la señal de muerte. Que suba aumenta la sensibilidad a morir.",
    "TRAIL R2/DR5": "Receptor de muerte celular muy estudiado como blanco terapéutico.",
    "Fas/TNFRSF6/CD95": "Receptor en la superficie de la célula que recibe órdenes de morir.",
    "HSP70": "Proteína protectora ante estrés. Que suba puede indicar resistencia al tratamiento.",
    "HSP27": "Bloquea la salida del citocromo c. Es un protector contra la muerte celular.",
    "p21/CIP1/CDKN1A": "Detiene la división celular. Su efecto depende del contexto (puede proteger o matar).",
    "Phospho-p53 (S15)": "Versión activa del 'guardián del genoma'. Cuando aparece, ordena a la célula morir.",
    "Phospho-p53 (S46)": "Versión de p53 que activa específicamente genes que matan a la célula.",
    "HIF-1a": "Aparece cuando la célula tiene poco oxígeno. En tumores ayuda a sobrevivir.",
    "Catalase": "Protege a la célula del daño por radicales libres. Que baje sensibiliza a morir.",
}


class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Proteome Profiler ARY009 — Semillero ITM")
        self.geometry("1100x820")
        self.minsize(900, 700)

        self._ruta = None
        self._cuant_lista: list = []        # 4 cuantificaciones individuales (una por strip)
        self._cuant_ctrl_avg: dict = {}     # promedio de strips 1-2
        self._cuant_trat_avg: dict = {}     # promedio de strips 3-4
        self._reporte: dict = {}            # reporte fold change con barras de error
        self._analizador = None             # referencia al analizador de grilla

        # Estado para modo de selección manual de referencias
        self._manual_strip_idx = None       # índice de la tirilla en modo manual
        self._manual_clicks: list = []      # lista de clicks (x, y) en modo manual
        self._manual_status_lbls: list = [None] * 4  # labels de estado por sub-tab
        self._ax_imgs: list = [None] * 4    # referencias a los ejes de imagen por sub-tab
        self._click_markers: list = []      # artistas matplotlib de los clicks actuales
        self._resumen_calidad_strip = (0, 0, 0)  # (alta, media, baja) de la tirilla actual

        self._build_ui()

    # ------------------------------------------------------------------ UI
    def _build_ui(self):
        ctk.CTkLabel(self, text="Analizador Proteome Profiler ARY009",
                     font=ctk.CTkFont(size=20, weight="bold")).pack(pady=(14, 2))
        ctk.CTkLabel(self, text="Semillero Biomédica · Detección de Cáncer · ITM",
                     font=ctk.CTkFont(size=11), text_color="gray").pack(pady=(0, 10))

        frame_cfg = ctk.CTkFrame(self)
        frame_cfg.pack(fill="x", padx=28, pady=4)
        frame_cfg.grid_columnconfigure((0, 1, 2), weight=1)

        self.btn_cargar = ctk.CTkButton(frame_cfg, text="Cargar imagen del escáner",
                                        command=self._cargar)
        self.btn_cargar.grid(row=0, column=0, padx=8, pady=12, sticky="ew")

        self.lbl_archivo = ctk.CTkLabel(frame_cfg, text="Sin imagen cargada",
                                        text_color="gray", wraplength=380, anchor="w")
        self.lbl_archivo.grid(row=0, column=1, columnspan=2, padx=8, pady=12, sticky="w")

        # Aviso sobre el diseño del kit ARY009
        ctk.CTkLabel(
            frame_cfg,
            text=("Esquema del kit ARY009: la imagen contiene 4 tirillas. "
                  "Las primeras 2 son Control y las últimas 2 son Tratamiento "
                  "(2 réplicas internas por condición)."),
            font=ctk.CTkFont(size=11),
            text_color="#9ad",
            wraplength=900, justify="left",
        ).grid(row=1, column=0, columnspan=3, padx=8, pady=(0, 8), sticky="w")

        self.btn_analizar = ctk.CTkButton(
            self, text="Analizar Membranas",
            fg_color="#28a745", hover_color="#1e7e34",
            font=ctk.CTkFont(size=15, weight="bold"),
            height=44, command=self._analizar)
        self.btn_analizar.pack(pady=12)

        # ── Tabview con 3 pestañas ──
        self.tabs = ctk.CTkTabview(self, height=450)
        self.tabs.pack(fill="both", expand=True, padx=28, pady=(0, 8))
        self.tabs.add("Reporte")
        self.tabs.add("Interpretación")
        self.tabs.add("Gráficos")
        self.tabs.add("Mapa Visual")

        # Tab 1: Reporte numérico
        self.textbox = ctk.CTkTextbox(
            self.tabs.tab("Reporte"),
            font=ctk.CTkFont(family="Consolas", size=10))
        self.textbox.pack(fill="both", expand=True, padx=4, pady=4)

        # Tab 2: Interpretación biológica
        self.txt_interp = ctk.CTkTextbox(
            self.tabs.tab("Interpretación"),
            font=ctk.CTkFont(family="Segoe UI", size=12),
            wrap="word")
        self.txt_interp.pack(fill="both", expand=True, padx=4, pady=4)
        self.txt_interp.insert("1.0",
            "  Ejecute un análisis para ver la interpretación biológica de los resultados.")

        # Tab 3: Gráficos (contenedor para el canvas matplotlib)
        self.frame_graf = ctk.CTkFrame(self.tabs.tab("Gráficos"), fg_color="transparent")
        self.frame_graf.pack(fill="both", expand=True, padx=4, pady=4)
        self._canvas_mpl = None
        self._lbl_graf_placeholder = ctk.CTkLabel(
            self.frame_graf, text="Ejecute un análisis para visualizar los gráficos.",
            text_color="gray")
        self._lbl_graf_placeholder.pack(expand=True)

        # Tab 4: Mapa Visual (detección de proteínas en la membrana)
        # Contiene un sub-tabview con una pestaña por tirilla individual.
        self.frame_mapa = ctk.CTkFrame(self.tabs.tab("Mapa Visual"), fg_color="transparent")
        self.frame_mapa.pack(fill="both", expand=True, padx=4, pady=4)
        self.tabs_mapa = ctk.CTkTabview(self.frame_mapa)
        self.tabs_mapa.pack(fill="both", expand=True)
        self._sub_tabs_mapa = [
            "Tirilla 1 (Control rep. 1)",
            "Tirilla 2 (Control rep. 2)",
            "Tirilla 3 (Tratam. rep. 1)",
            "Tirilla 4 (Tratam. rep. 2)",
            "Guía de Proteínas",
        ]
        for nombre in self._sub_tabs_mapa:
            self.tabs_mapa.add(nombre)
        self._canvas_sub_mapas = [None] * 4

        # Cargar y mostrar la imagen estática en la pestaña "Guía de Proteínas"
        from PIL import Image
        frame_guia = ctk.CTkScrollableFrame(self.tabs_mapa.tab("Guía de Proteínas"), fg_color="transparent")
        frame_guia.pack(fill="both", expand=True)
        mapa_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "mapa_referencia_ary009.png")
        try:
            if os.path.exists(mapa_path):
                img_ref = Image.open(mapa_path)
                ctk_img = ctk.CTkImage(light_image=img_ref, dark_image=img_ref, size=(600, 900))
                lbl_ref = ctk.CTkLabel(frame_guia, image=ctk_img, text="")
                lbl_ref.pack(expand=True, pady=10)
            else:
                lbl_ref = ctk.CTkLabel(frame_guia, text="Generando mapa de referencia...", text_color="gray")
                lbl_ref.pack(expand=True)
        except Exception as e:
            lbl_ref = ctk.CTkLabel(frame_guia, text=f"Error cargando mapa: {e}", text_color="gray")
            lbl_ref.pack(expand=True)
        ctk.CTkLabel(
            self.tabs_mapa.tab(self._sub_tabs_mapa[0]),
            text="Ejecute un análisis para ver el mapa visual de cada tirilla.",
            text_color="gray").pack(expand=True)

        self.btn_export = ctk.CTkButton(
            self, text="Exportar resultados a Excel",
            fg_color="#0d6efd", hover_color="#0a58ca",
            command=self._exportar, state="disabled")
        self.btn_export.pack(pady=(0, 14))

    # ------------------------------------------------------------------ carga
    def _cargar(self):
        ruta = filedialog.askopenfilename(
            title="Seleccionar imagen del escáner",
            filetypes=[
                ("Imágenes soportadas",
                 "*.jpg *.jpeg *.png *.bmp *.tif *.tiff *.scn *.svs *.ndpi"),
                ("Todos los archivos", "*.*"),
            ])
        if ruta:
            self._ruta = ruta
            self.lbl_archivo.configure(
                text=os.path.basename(ruta), text_color=("black", "white"))

    # ------------------------------------------------------------------ análisis
    def _analizar(self):
        if not self._ruta:
            messagebox.showwarning("Falta imagen", "Cargue una imagen primero.")
            return
        self.btn_analizar.configure(state="disabled", text="Analizando…")
        self.textbox.delete("1.0", ctk.END)
        self._log("Iniciando análisis…\n\n")

        threading.Thread(target=self._run_analisis, daemon=True).start()

    def _run_analisis(self):
        try:
            dir_img = os.path.dirname(self._ruta)
            ruta_debug = os.path.join(dir_img, "debug_grilla.png")

            self._log("→ Separando 4 membranas y anclando grilla en Reference Spots…\n")
            self._analizador = AnalizadorGrillaARY009(radio_muestreo=8)
            intensidades_por_strip = self._analizador.analizar_imagen(
                self._ruta, n_strips=4, ruta_debug=ruta_debug)

            self._cuant_lista = [self._analizador.cuantificar(inten)
                                  for inten in intensidades_por_strip]

            if len(self._cuant_lista) < 4:
                self._log(
                    f"[!] Solo se detectaron {len(self._cuant_lista)} membranas. "
                    "El kit ARY009 requiere 4 (2 control + 2 tratamiento). "
                    "Verifique que la imagen contenga las 4 tirillas completas.\n")
                return

            # Esquema oficial ARY009: strips 1-2 = control, strips 3-4 = tratamiento
            cuant_ctrl_strips = self._cuant_lista[0:2]
            cuant_trat_strips = self._cuant_lista[2:4]

            n_c1 = len(cuant_ctrl_strips[0]['proteinas'])
            n_c2 = len(cuant_ctrl_strips[1]['proteinas'])
            n_t1 = len(cuant_trat_strips[0]['proteinas'])
            n_t2 = len(cuant_trat_strips[1]['proteinas'])
            self._log(f"   Strip 1 (Control rep1):     {n_c1} proteínas\n")
            self._log(f"   Strip 2 (Control rep2):     {n_c2} proteínas\n")
            self._log(f"   Strip 3 (Tratamiento rep1): {n_t1} proteínas\n")
            self._log(f"   Strip 4 (Tratamiento rep2): {n_t2} proteínas\n\n")
            self._log(f"   Imagen debug guardada: {ruta_debug}\n\n")

            self._log("→ Promediando réplicas internas (strips 1-2 y 3-4)…\n")
            self._cuant_ctrl_avg = promediar_cuantificaciones(cuant_ctrl_strips)
            self._cuant_trat_avg = promediar_cuantificaciones(cuant_trat_strips)

            self._log("→ Calculando Fold Change con propagación de error…\n\n")
            self._reporte = calcular_fold_change_con_error(
                self._cuant_ctrl_avg,
                self._cuant_trat_avg,
            )

            self._mostrar_reporte()
            self.after(0, self._mostrar_interpretacion)
            self.after(0, self._mostrar_graficos)
            self.after(0, self._mostrar_mapa_visual)
            self.after(0, lambda: self.btn_export.configure(state="normal"))

        except Exception:
            self._log(f"\n[ERROR]\n{traceback.format_exc()}\n")
        finally:
            self.after(0, lambda: self.btn_analizar.configure(
                state="normal", text="Analizar Membranas"))

    # ------------------------------------------------------------------ display: REPORTE
    def _mostrar_reporte(self):
        r = self._reporte
        cuant_c = self._cuant_ctrl_avg
        cuant_t = self._cuant_trat_avg

        lineas = [
            "=" * 78,
            "  REPORTE DE RESULTADOS",
            "  Control = promedio de tirillas 1-2  |  Tratamiento = promedio de tirillas 3-4",
            "=" * 78,
            "",
            "  CÓMO LEER LA TABLA:",
            "    • 'Control' y 'Tratam.': intensidad de la proteína en cada condición.",
            "    • '±': qué tan parecidas son las dos réplicas internas.",
            "    • 'Veces': cuántas veces cambió la proteína (1.0 = igual).",
            "    • '± err': margen de error de ese cambio.",
            "    • '¿Qué pasó?': resumen en español de lo que ocurrió.",
            "",
            f"  Referencia control:      {cuant_c.get('referencia', 0):.1f}  "
            f"(útil: {cuant_c.get('referencia_neta', 0):.1f})",
            f"  Referencia tratamiento:  {cuant_t.get('referencia', 0):.1f}  "
            f"(útil: {cuant_t.get('referencia_neta', 0):.1f})",
            f"  Fondo PBS control:       {cuant_c.get('pbs', 0):.1f}    "
            f"Fondo PBS tratamiento:  {cuant_t.get('pbs', 0):.1f}",
            "",
        ]

        por_tipo = {
            "ACTIVADORAS DE MUERTE CELULAR (pro-apoptóticas)": [],
            "BLOQUEADORAS DE MUERTE CELULAR (anti-apoptóticas)": [],
            "OTRAS PROTEÍNAS": [],
        }
        for prot, datos in sorted(r.items()):
            g = ("ACTIVADORAS DE MUERTE CELULAR (pro-apoptóticas)" if datos['tipo'] == 'pro-apoptótica'
                 else "BLOQUEADORAS DE MUERTE CELULAR (anti-apoptóticas)" if datos['tipo'] == 'anti-apoptótica'
                 else "OTRAS PROTEÍNAS")
            por_tipo[g].append((prot, datos))

        for grupo, items in por_tipo.items():
            if not items:
                continue
            lineas.append(f"── {grupo} " + "─" * max(1, 78 - len(grupo) - 4))
            lineas.append(
                f"  {'Proteína':<26} {'Control':>8} {'±':>5} "
                f"{'Tratam.':>8} {'±':>5} {'Veces':>7} {'± err':>6}  ¿Qué pasó?")
            lineas.append("  " + "─" * 76)
            items_ord = sorted(
                items,
                key=lambda x: (x[1]['fold_change']
                               if x[1]['fold_change'] != float('inf') else 999),
                reverse=True)
            for prot, d in items_ord:
                fold = d['fold_change']
                fold_s = "∞" if fold == float('inf') else f"{fold:.3f}"
                err_s  = "" if fold == float('inf') or fold == 0.0 else f"{d.get('fold_err', 0):.3f}"
                lineas.append(
                    f"  {prot:<26} {d['ctrl_norm']:>8.4f} ±{d.get('ctrl_std', 0):>4.3f} "
                    f"{d['trat_norm']:>8.4f} ±{d.get('trat_std', 0):>4.3f} "
                    f"{fold_s:>7} {err_s:>6}  {d['estado']}")
            lineas.append("")

        ups = [(p, d['fold_change']) for p, d in r.items()
               if d['fold_change'] not in (float('inf'), 0.0) and d['fold_change'] >= 1.5]
        downs = [(p, d['fold_change']) for p, d in r.items()
                 if d['fold_change'] not in (float('inf'), 0.0) and d['fold_change'] <= 0.67]
        ups.sort(key=lambda x: x[1], reverse=True)
        downs.sort(key=lambda x: x[1])

        lineas += ["=" * 78, "  PROTEÍNAS QUE MÁS SUBIERON (al menos 1.5 veces)"]
        for p, f in ups[:8]:
            lineas.append(f"    {p:<28} {f:.3f} veces")
        lineas += ["", "  PROTEÍNAS QUE MÁS BAJARON (al menos 33% menos)"]
        for p, f in downs[:8]:
            lineas.append(f"    {p:<28} {f:.3f} veces")
        lineas.append("=" * 78)

        texto = "\n".join(lineas)
        self.after(0, lambda: self.textbox.insert(ctk.END, texto + "\n"))

    # ------------------------------------------------------------------ display: INTERPRETACIÓN
    def _mostrar_interpretacion(self):
        r = self._reporte
        if not r:
            return

        # Conteos por categoría y dirección
        conteos = {
            'pro': {'up': [], 'down': [], 'denovo': [], 'silenciada': [], 'sin_cambio': []},
            'anti': {'up': [], 'down': [], 'denovo': [], 'silenciada': [], 'sin_cambio': []},
        }
        for prot, d in r.items():
            tipo = d['tipo']
            if tipo == 'pro-apoptótica':
                k = 'pro'
            elif tipo == 'anti-apoptótica':
                k = 'anti'
            else:
                continue
            est = d['estado']
            if 'APARECIÓ' in est:
                conteos[k]['denovo'].append(prot)
            elif 'DESAPARECIÓ' in est:
                conteos[k]['silenciada'].append(prot)
            elif 'AUMENTÓ' in est:
                conteos[k]['up'].append((prot, d['fold_change']))
            elif 'DISMINUYÓ' in est:
                conteos[k]['down'].append((prot, d['fold_change']))
            else:
                conteos[k]['sin_cambio'].append(prot)

        # Score apoptótico: pro↑ y anti↓ son favorables a apoptosis
        score = (len(conteos['pro']['up']) + len(conteos['pro']['denovo'])
                 + len(conteos['anti']['down']) + len(conteos['anti']['silenciada'])
                 - len(conteos['pro']['down']) - len(conteos['pro']['silenciada'])
                 - len(conteos['anti']['up']) - len(conteos['anti']['denovo']))

        if score >= 5:
            veredicto = "⬆️  INDUCCIÓN DE APOPTOSIS PROBABLE"
            color_v = "El balance favorece claramente la muerte celular programada."
        elif score >= 2:
            veredicto = "↗  TENDENCIA PRO-APOPTÓTICA"
            color_v = "Hay señales pro-apoptóticas dominantes pero no contundentes."
        elif score <= -5:
            veredicto = "⬇️  RESISTENCIA / SUPERVIVENCIA DOMINANTE"
            color_v = "El tratamiento parece favorecer la supervivencia celular."
        elif score <= -2:
            veredicto = "↘  TENDENCIA ANTI-APOPTÓTICA"
            color_v = "Hay señales de resistencia pero no contundentes."
        else:
            veredicto = "—  RESPUESTA AMBIGUA O SIN CAMBIO"
            color_v = "El balance pro/anti está equilibrado; el tratamiento no muestra efecto claro."

        L = []
        L.append("═" * 70)
        L.append(f"  INTERPRETACIÓN DEL EXPERIMENTO")
        L.append(f"  Control: tirillas 1-2 (promedio)  →  Tratamiento: tirillas 3-4 (promedio)")
        L.append("═" * 70)
        L.append("")
        L.append("  📖 PARA ENTENDER ESTE REPORTE")
        L.append("  ──────────────────────────────")
        L.append("  • APOPTOSIS = muerte celular programada (lo que buscamos en cáncer).")
        L.append("  • Proteína 'pro-apoptótica' = AYUDA a matar la célula cancerosa.")
        L.append("  • Proteína 'anti-apoptótica' = PROTEGE a la célula cancerosa.")
        L.append("  • Un buen tratamiento debería SUBIR las pro-apoptóticas y BAJAR las anti.")
        L.append("")
        L.append("  📊 SIGNIFICADO DE CADA ESTADO")
        L.append("  ──────────────────────────────")
        L.append("  ▲ AUMENTÓ      → la proteína subió más de 1.5 veces.")
        L.append("  ▼ DISMINUYÓ    → la proteína bajó por debajo del 67%.")
        L.append("  ▲ APARECIÓ     → no estaba en el control y aparece en el tratamiento.")
        L.append("  ▼ DESAPARECIÓ  → estaba en el control y se va con el tratamiento.")
        L.append("  ─ SIN CAMBIO   → la proteína está igual o el cambio es muy pequeño.")
        L.append("")
        L.append(f"  🔎 VEREDICTO: {veredicto}")
        L.append(f"  {color_v}")
        L.append(f"  (Puntaje apoptótico = {score:+d})")
        L.append("")
        L.append("─" * 70)
        L.append("  📊 RESUMEN: ¿CUÁNTAS PROTEÍNAS CAMBIARON?")
        L.append("─" * 70)
        L.append(f"  Activadoras de muerte celular (pro-apoptóticas):")
        L.append(f"     ▲ Subieron: {len(conteos['pro']['up']):>2}     "
                 f"▼ Bajaron: {len(conteos['pro']['down']):>2}     "
                 f"▲ Aparecieron: {len(conteos['pro']['denovo']):>2}     "
                 f"▼ Desaparecieron: {len(conteos['pro']['silenciada']):>2}     "
                 f"─ Sin cambio: {len(conteos['pro']['sin_cambio']):>2}")
        L.append(f"  Bloqueadoras de muerte celular (anti-apoptóticas):")
        L.append(f"     ▲ Subieron: {len(conteos['anti']['up']):>2}     "
                 f"▼ Bajaron: {len(conteos['anti']['down']):>2}     "
                 f"▲ Aparecieron: {len(conteos['anti']['denovo']):>2}     "
                 f"▼ Desaparecieron: {len(conteos['anti']['silenciada']):>2}     "
                 f"─ Sin cambio: {len(conteos['anti']['sin_cambio']):>2}")
        L.append("")

        # Sección: cómo leer el score
        L.append("─" * 70)
        L.append("  ¿CÓMO SE CALCULA EL PUNTAJE APOPTÓTICO?")
        L.append("─" * 70)
        L.append("  Suma + 1 punto cada vez que ocurre algo BUENO contra el cáncer:")
        L.append("    · Proteína activadora de muerte SUBE o APARECE.")
        L.append("    · Proteína bloqueadora de muerte BAJA o DESAPARECE.")
        L.append("  Resta - 1 punto cada vez que ocurre algo MALO:")
        L.append("    · Proteína activadora de muerte BAJA o DESAPARECE.")
        L.append("    · Proteína bloqueadora de muerte SUBE o APARECE.")
        L.append("  Puntaje alto positivo = el tratamiento parece estar funcionando.")
        L.append("")

        # Hallazgos clave
        L.append("─" * 70)
        L.append("  🔬 HALLAZGOS PRINCIPALES (proteínas más importantes que cambiaron)")
        L.append("─" * 70)

        relevantes = []
        for prot, d in r.items():
            if prot not in PROTEINAS_CLAVE:
                continue
            est = d['estado']
            if 'SIN CAMBIO' in est:
                continue
            relevantes.append((prot, d, est))

        if not relevantes:
            L.append("  (Ninguna proteína clave mostró cambio relevante)")
        else:
            # Ordenar por magnitud absoluta del cambio
            def magnitud(item):
                f = item[1]['fold_change']
                if f == float('inf'):
                    return 999
                if f == 0:
                    return 999
                return max(f, 1 / f) if f > 0 else 999
            relevantes.sort(key=magnitud, reverse=True)

            for prot, d, est in relevantes[:10]:
                fold = d['fold_change']
                if fold == float('inf'):
                    fold_s = "APARECIÓ"
                elif fold == 0:
                    fold_s = "DESAPARECIÓ"
                else:
                    fold_s = f"{fold:.2f}×"
                tipo_legible = ("activadora de muerte celular" if d['tipo'] == 'pro-apoptótica'
                                else "bloqueadora de muerte celular" if d['tipo'] == 'anti-apoptótica'
                                else "otra")
                L.append(f"  • {prot}  [{fold_s}]  ({tipo_legible})")
                L.append(f"      {PROTEINAS_CLAVE[prot]}")
                L.append("")

        # Razón Bax/Bcl-2 (clásico indicador de apoptosis)
        if 'Bax' in r and 'Bcl-2' in r:
            bax_t = r['Bax']['trat_norm']
            bcl_t = r['Bcl-2']['trat_norm']
            bax_c = r['Bax']['ctrl_norm']
            bcl_c = r['Bcl-2']['ctrl_norm']
            if bcl_c > 0.001 and bcl_t > 0.001:
                ratio_c = bax_c / bcl_c
                ratio_t = bax_t / bcl_t
                cambio = ratio_t / ratio_c if ratio_c > 0.001 else float('inf')
                L.append("─" * 70)
                L.append("  ⚖️  BALANCE BAX / BCL-2  (uno de los indicadores más usados)")
                L.append("─" * 70)
                L.append("  Bax es la proteína que ABRE la mitocondria para matar la célula.")
                L.append("  Bcl-2 es la que la BLOQUEA. Comparar las dos es muy informativo.")
                L.append("")
                L.append(f"  En el control:      {ratio_c:.3f}")
                L.append(f"  En el tratamiento:  {ratio_t:.3f}")
                if isinstance(cambio, float) and cambio != float('inf'):
                    L.append(f"  Cambio:             {cambio:.2f} veces")
                if ratio_t > ratio_c * 1.3:
                    L.append("  → Bax domina sobre Bcl-2: la balanza se inclina hacia muerte ✓")
                elif ratio_t < ratio_c * 0.77:
                    L.append("  → Bcl-2 domina sobre Bax: la balanza se inclina hacia supervivencia")
                else:
                    L.append("  → Sin cambio importante en el balance.")
                L.append("")

        L.append("═" * 70)
        L.append("  ⚠️  IMPORTANTE")
        L.append("  Esta interpretación es ORIENTATIVA. Para tener certeza científica")
        L.append("  hay que repetir el experimento varias veces (mínimo 3) y confirmar")
        L.append("  con otros ensayos como Western blot.")
        L.append("")
        L.append("  ⚠️  NOTA SOBRE VALORES p")
        L.append("  Con solo 2 réplicas internas por condición, los p-valores reportados")
        L.append("  son ESTIMACIONES con muy bajo poder estadístico (df≈2). NO deben")
        L.append("  usarse como criterio definitivo. Para significancia estadística real")
        L.append("  se necesitan mínimo 3 réplicas biológicas independientes.")
        L.append("═" * 70)

        texto = "\n".join(L)
        self.txt_interp.delete("1.0", ctk.END)
        self.txt_interp.insert("1.0", texto)

    # ------------------------------------------------------------------ display: GRÁFICOS
    def _mostrar_graficos(self):
        """
        Recrea el formato de los paneles del paper:
          - Panel A: Pro-apoptóticas y caspasas (barras horizontales con error bars)
          - Panel C: Anti-apoptóticas (barras horizontales con error bars)
        Eje X: Fold Change relativo (Tratamiento / Control)
        """
        r = self._reporte
        if not r:
            return

        # Limpiar canvas previo
        if self._canvas_mpl is not None:
            self._canvas_mpl.get_tk_widget().destroy()
            self._canvas_mpl = None
        self._lbl_graf_placeholder.pack_forget()
        for w in self.frame_graf.winfo_children():
            w.destroy()

        fig = Figure(figsize=(11, 8), dpi=90, facecolor="#212121")
        gs = fig.add_gridspec(2, 1, hspace=0.45,
                              left=0.22, right=0.96, top=0.94, bottom=0.07)

        ax_pro  = fig.add_subplot(gs[0, 0])
        ax_anti = fig.add_subplot(gs[1, 0])

        for ax in (ax_pro, ax_anti):
            ax.set_facecolor("#2a2a2a")
            for s in ax.spines.values():
                s.set_color("#888")
            ax.tick_params(colors="#ddd", labelsize=9)
            ax.title.set_color("#fff")
            ax.xaxis.label.set_color("#ddd")
            ax.yaxis.label.set_color("#ddd")

        def _fold_relativo(d):
            """Convierte fold_change → fold relativo centrado en 0.
            Bajada (fold < 1) → negativo (-1 a 0); Subida (fold > 1) → positivo (0 a +N).
            Sirve para replicar el formato del paper (eje X: -1 a +3)."""
            f = d['fold_change']
            if f == float('inf'):
                return 3.0  # cap visual
            if f == 0.0:
                return -1.0
            return f - 1.0

        def _err_relativo(d):
            f = d['fold_change']
            if f in (float('inf'), 0.0):
                return 0.0
            return d.get('fold_err', 0.0)

        def _dibujar_panel(ax, lista_proteinas, titulo, color_barra):
            """
            Dibuja un panel del paper:
            - Barras horizontales centradas en 0
            - Error bars
            - Etiqueta de p-valor (si hay réplicas suficientes)
            """
            datos = []
            for prot in lista_proteinas:
                if prot not in r:
                    continue
                d = r[prot]
                fr = _fold_relativo(d)
                er = _err_relativo(d)
                datos.append((prot, fr, er, d))

            if not datos:
                ax.text(0.5, 0.5, "Sin datos para esta categoría",
                        transform=ax.transAxes, ha="center", color="#888")
                ax.set_title(titulo, fontsize=12, weight="bold")
                return

            # Mantener orden bio (no orden alfabético) — como en el paper
            nombres = [d[0] for d in datos]
            valores = [d[1] for d in datos]
            errores = [d[2] for d in datos]

            y_pos = np.arange(len(nombres))
            ax.barh(y_pos, valores, xerr=errores,
                    color=color_barra, edgecolor="#222",
                    error_kw={'ecolor': '#ddd', 'lw': 1, 'capsize': 3})

            ax.set_yticks(y_pos)
            ax.set_yticklabels(nombres, fontsize=9)
            ax.invert_yaxis()  # primer ítem arriba (formato paper)
            ax.axvline(0, color="#fff", lw=0.8)
            ax.set_xlim(-1.2, 3.2)
            ax.set_xlabel("Cambio respecto al control   (← bajó      |      subió →)",
                          fontsize=10)
            ax.set_title(titulo, fontsize=12, weight="bold")
            ax.grid(axis="x", alpha=0.2)

            # Anotar p-valores si las réplicas internas dan suficiente info
            for i, (prot, val, err, d) in enumerate(datos):
                if d.get('n_ctrl', 1) >= 2 and d.get('n_trat', 1) >= 2:
                    # Test t aproximado con 2 réplicas (poco potente, indicativo)
                    p_aprox = self._p_value_aprox(d)
                    if p_aprox is not None and p_aprox < 0.05:
                        x_text = val + (0.15 if val >= 0 else -0.15)
                        ax.text(x_text, i, f"p={p_aprox:.4f}",
                                fontsize=7, color="#ffd54f",
                                ha="left" if val >= 0 else "right",
                                va="center")

        # Panel pro-apoptóticas (orden similar al paper)
        orden_pro = [
            "Phospho-Rad17 (S635)",
            "Pro-Caspase-3", "Cleaved Caspase-3",
            "Cytochrome c",
            "Phospho-p53 (S392)", "Phospho-p53 (S46)", "Phospho-p53 (S15)",
            "Bad", "Bax",
            "TRAIL R2/DR5", "TRAIL R1/DR4",
            "SMAC/Diablo",
        ]
        _dibujar_panel(ax_pro, orden_pro,
                       "Proteínas que ACTIVAN la muerte celular  (queremos que SUBAN)",
                       "#4d6fb3")

        # Panel anti-apoptóticas (orden similar al paper)
        orden_anti = [
            "Livin", "cIAP-2", "cIAP-1",
            "Survivin", "XIAP",
            "Bcl-x", "Bcl-2",
        ]
        _dibujar_panel(ax_anti, orden_anti,
                       "Proteínas que BLOQUEAN la muerte celular  (queremos que BAJEN)",
                       "#5b8def")

        canvas = FigureCanvasTkAgg(fig, master=self.frame_graf)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True)
        self._canvas_mpl = canvas

    # ------------------------------------------------------------------ display: MAPA VISUAL
    def _mostrar_mapa_visual(self):
        """
        Genera UNA pestaña por tirilla, mostrando en cada una:
          - Lado izquierdo: imagen real de la tirilla con cuadros marcando las
            áreas que el programa muestreó (lo que el programa "ve").
          - Lado derecho: gráfico de intensidades detectadas en cada proteína,
            agrupadas por categoría y ordenadas de mayor a menor.
        """
        if self._analizador is None:
            return

        datos = self._analizador.obtener_datos_mapa_visual()
        if not datos:
            return

        def color_tipo(nombre):
            if nombre == 'Reference':
                return '#FFD700'
            if nombre == 'PBS':
                return '#00CED1'
            if nombre in PRO_APOPTOTICAS:
                return '#FF4444'
            if nombre in ANTI_APOPTOTICAS:
                return '#4488FF'
            return '#88FF88'

        from matplotlib.patches import Rectangle
        from matplotlib.lines import Line2D

        for i in range(min(len(datos), len(self._sub_tabs_mapa))):
            d = datos[i]
            nombre_tab = self._sub_tabs_mapa[i]
            tab_frame = self.tabs_mapa.tab(nombre_tab)

            # Limpiar contenido previo
            if self._canvas_sub_mapas[i] is not None:
                self._canvas_sub_mapas[i].get_tk_widget().destroy()
                self._canvas_sub_mapas[i] = None
            for w in tab_frame.winfo_children():
                w.destroy()

            # ── Barra superior con botón de modo manual ──
            top_bar = ctk.CTkFrame(tab_frame, fg_color="transparent")
            top_bar.pack(fill="x", padx=8, pady=(8, 4))

            ctk.CTkLabel(
                top_bar,
                text="¿Las referencias están mal? →",
                text_color="#bbb", font=ctk.CTkFont(size=11),
            ).pack(side="left", padx=(0, 8))

            ctk.CTkButton(
                top_bar,
                text="Ajustar referencias manualmente",
                width=240, height=28,
                fg_color="#fd7e14", hover_color="#d96400",
                font=ctk.CTkFont(size=11, weight="bold"),
                command=lambda idx=i: self._iniciar_modo_manual(idx),
            ).pack(side="left")

            self._manual_status_lbls[i] = ctk.CTkLabel(
                top_bar,
                text="",
                text_color="#ffd54f",
                font=ctk.CTkFont(size=11, weight="bold"),
            )
            self._manual_status_lbls[i].pack(side="left", padx=(12, 0))

            # Aviso de auto-detección fallida (la imagen igual se muestra abajo)
            if d is None or d.get('auto_failed') or not d.get('posiciones'):
                aviso = ctk.CTkFrame(tab_frame, fg_color="#3a2e15", corner_radius=6)
                aviso.pack(fill="x", padx=8, pady=(4, 0))
                ctk.CTkLabel(
                    aviso,
                    text=("⚠ La detección automática falló para esta tirilla.  "
                          "Hacé click en 'Ajustar referencias manualmente' "
                          "y luego clickeá los 3 Reference Spots reales sobre la imagen."),
                    text_color="#ffb84d",
                    font=ctk.CTkFont(size=11, weight="bold"),
                    justify="left",
                ).pack(padx=10, pady=6, anchor="w")

            if d is None:
                # Sin imagen disponible — no debería pasar pero por seguridad
                ctk.CTkLabel(
                    tab_frame,
                    text="No hay imagen disponible para esta tirilla.",
                    text_color="gray",
                ).pack(expand=True)
                continue

            cuant = self._cuant_lista[i] if i < len(self._cuant_lista) else None
            proteinas = cuant.get('proteinas', {}) if cuant else {}

            # ── Crear figura con 1 panel (Solo imagen) ──
            fig = Figure(figsize=(14, 8.5), dpi=90, facecolor='#1a1a2e')
            gs = fig.add_gridspec(
                1, 1,
                left=0.05, right=0.95, top=0.95, bottom=0.05,
            )

            # ═══════════════════════════════════════════════════════════
            # PANEL IZQUIERDO: tirilla con áreas muestreadas
            # ═══════════════════════════════════════════════════════════
            ax_img = fig.add_subplot(gs[0])
            self._ax_imgs[i] = ax_img  # guardar para captura de clicks manuales
            strip_img = d['strip_img']
            h_img, w_img = strip_img.shape[:2]

            ax_img.imshow(strip_img, cmap='gray', aspect='equal', zorder=1)
            ax_img.set_facecolor('#1a1a2e')

            analizador = self._analizador
            refs = d.get('refs_puntos', [None, None, None])

            # Helper: desempaqueta (px, py, proteina) o (px, py, proteina, calidad)
            def _unpack(val):
                if len(val) >= 4:
                    return val[0], val[1], val[2], float(val[3])
                return val[0], val[1], val[2], 1.0

            # Contadores de calidad para el resumen
            n_alta = n_media = n_baja = 0

            # Dibujar cada área muestreada con estilo según calidad
            for (col, fila), val in d['posiciones'].items():
                px, py, proteina, calidad = _unpack(val)
                c = color_tipo(proteina)
                radio = analizador.get_radio_dinamico(refs[0], refs[2]) if None not in refs else analizador.radio

                # Estilo según calidad:
                #   alta (>=0.6): borde sólido grueso, alpha 0.95
                #   media (0.3-0.6): borde sólido normal, alpha 0.7
                #   baja (<0.3): borde punteado, alpha 0.35 (cuadro vacío)
                if calidad >= 0.6:
                    estilo = '-'
                    grosor = 1.2
                    alpha = 0.95
                    n_alta += 1
                elif calidad >= 0.3:
                    estilo = '-'
                    grosor = 0.9
                    alpha = 0.70
                    n_media += 1
                else:
                    estilo = ':'
                    grosor = 0.7
                    alpha = 0.35
                    n_baja += 1

                rect = Rectangle(
                    (px - radio, py - radio), radio * 2, radio * 2,
                    linewidth=grosor, edgecolor=c, facecolor='none',
                    alpha=alpha, linestyle=estilo, zorder=3,
                )
                ax_img.add_patch(rect)

            # Guardar resumen de calidad para mostrar después
            self._resumen_calidad_strip = (n_alta, n_media, n_baja)

            # Letras de columna (A-E) sobre la imagen
            col_xs = {}
            for (col, fila), val in d['posiciones'].items():
                px = val[0]
                col_xs.setdefault(col, []).append(px)
            for col_letra, xs in col_xs.items():
                x_avg = float(np.mean(xs))
                ax_img.text(x_avg, -10, col_letra,
                            ha='center', va='bottom',
                            fontsize=10, color='#ddd', weight='bold', zorder=5)

            # Recortar el eje Y para mostrar solo la zona de las manchas
            y_coords = [val[1] for val in d['posiciones'].values()]
            if y_coords:
                min_y = max(0, min(y_coords) - 30)
                max_y = min(h_img, max(y_coords) + 30)
                ax_img.set_ylim(max_y, min_y) # Invertido porque en imágenes Y crece hacia abajo
            else:
                ax_img.set_ylim(h_img + 5, -28)

            ax_img.set_xlim(-5, w_img + 5)

            # Resumen de calidad en el título
            n_alta, n_media, n_baja = self._resumen_calidad_strip
            n_total = n_alta + n_media + n_baja
            if n_total > 0:
                pct_ok = int(100 * n_alta / n_total)
                titulo_extra = (f"  •  {n_alta} cuadros sobre spots reales · "
                                f"{n_baja} sobre vacío  ({pct_ok}% calidad)")
            else:
                titulo_extra = ""

            ax_img.set_title(
                f"Lo que VE el programa{titulo_extra}",
                color='white', fontsize=10, weight='bold', pad=8,
            )
            ax_img.axis('off')

            # Leyenda extendida: colores + estilos de calidad
            leyenda = [
                Line2D([0], [0], marker='s', color='w', markerfacecolor='none',
                       markeredgecolor='#FFD700', markeredgewidth=1.5,
                       markersize=8, linestyle='None', label='Referencia'),
                Line2D([0], [0], marker='s', color='w', markerfacecolor='none',
                       markeredgecolor='#00CED1', markeredgewidth=1.5,
                       markersize=8, linestyle='None', label='PBS'),
                Line2D([0], [0], marker='s', color='w', markerfacecolor='none',
                       markeredgecolor='#FF4444', markeredgewidth=1.5,
                       markersize=8, linestyle='None', label='Activadora'),
                Line2D([0], [0], marker='s', color='w', markerfacecolor='none',
                       markeredgecolor='#4488FF', markeredgewidth=1.5,
                       markersize=8, linestyle='None', label='Bloqueadora'),
                Line2D([0], [0], marker='s', color='w', markerfacecolor='none',
                       markeredgecolor='#cccccc', markeredgewidth=1.4,
                       markersize=8, linestyle='None', label='─ sólido = sobre spot'),
                Line2D([0], [0], marker='s', color='w', markerfacecolor='none',
                       markeredgecolor='#888888', markeredgewidth=0.8,
                       markersize=8, linestyle='None', label='┄ punteado = sin spot'),
            ]
            ax_img.legend(handles=leyenda, loc='lower center',
                          bbox_to_anchor=(0.5, -0.08),
                          ncol=3, fontsize=7,
                          facecolor='#2a2a3e', edgecolor='#555',
                          labelcolor='white', framealpha=0.9)

            canvas = FigureCanvasTkAgg(fig, master=tab_frame)
            canvas.draw()
            canvas.get_tk_widget().pack(fill="both", expand=True, padx=8, pady=(0, 2))
            self._canvas_sub_mapas[i] = canvas

            # Agregar barra de herramientas de Matplotlib (permite Zoom y Paneo)
            toolbar_frame = ctk.CTkFrame(tab_frame, height=35, fg_color="transparent")
            toolbar_frame.pack(fill="x", padx=8, pady=(0, 8))
            toolbar = NavigationToolbar2Tk(canvas, toolbar_frame)
            toolbar.update()
            toolbar.pack(side="left")

            # Conectar evento de click para selección manual de referencias
            canvas.mpl_connect(
                'button_press_event',
                lambda event, idx=i: self._on_canvas_click(event, idx),
            )

    # ------------------------------------------------------------------ MODO MANUAL: selección de referencias
    def _iniciar_modo_manual(self, strip_idx):
        """
        Activa el modo de selección manual de referencias para una tirilla.
        El usuario debe hacer 4 clicks sobre los Reference Spots/PBS reales.
        """
        self._manual_strip_idx = strip_idx
        self._manual_clicks = []
        self._click_markers = []
        self._actualizar_status_manual(
            strip_idx,
            "Click 1: ESQ. SUP. DERECHA (A1/A2 - Derecha Arriba)",
        )

    def _actualizar_status_manual(self, strip_idx, texto, color="#ffd54f"):
        """Actualiza el label de estado del modo manual."""
        if 0 <= strip_idx < len(self._manual_status_lbls):
            lbl = self._manual_status_lbls[strip_idx]
            if lbl is not None:
                lbl.configure(text=texto, text_color=color)

    def _on_canvas_click(self, event, strip_idx):
        """
        Captura clicks sobre el canvas durante el modo manual.
        Solo responde si el modo manual está activo para esta tirilla
        Y el click cae sobre el panel de la imagen (no sobre el gráfico).
        """
        # Solo activo si esta tirilla está en modo manual
        if self._manual_strip_idx != strip_idx:
            return
        if event.xdata is None or event.ydata is None or event.inaxes is None:
            return
        # Verificar que el click cayó sobre el panel de la imagen
        if event.inaxes is not self._ax_imgs[strip_idx]:
            return

        x, y = float(event.xdata), float(event.ydata)
        self._manual_clicks.append((x, y))

        # Dibujar marcador en el click
        ax = self._ax_imgs[strip_idx]
        labels = ["A_sup", "E_sup", "A_inf", "D_inf"]
        idx_click = len(self._manual_clicks) - 1
        marker = ax.plot(
            x, y, 'X', color='#fd7e14', markersize=14,
            markeredgecolor='white', markeredgewidth=1.5, zorder=10,
        )[0]
        text = ax.annotate(
            labels[idx_click],
            xy=(x, y),
            xytext=(x + 8, y - 8),
            color='#fd7e14', fontsize=11, weight='bold',
            zorder=11,
        )
        self._click_markers.append(marker)
        self._click_markers.append(text)
        self._canvas_sub_mapas[strip_idx].draw_idle()

        n = len(self._manual_clicks)
        if n == 1:
            self._actualizar_status_manual(
                strip_idx, "Click 2: ESQ. SUP. IZQUIERDA (E1/E2)")
        elif n == 2:
            self._actualizar_status_manual(
                strip_idx, "Click 3: ESQ. INF. DERECHA (A23/A24)")
        elif n == 3:
            self._actualizar_status_manual(
                strip_idx, "Click 4: MANCHA INF. IZQUIERDA (PBS - D23/D24)")
        else:
            # 4 clicks completos → aplicar
            self._aplicar_refs_manuales(strip_idx)

    def _aplicar_refs_manuales(self, strip_idx):
        """
        Aplica las 4 referencias manuales: re-muestrea la tirilla,
        recalcula promedios y fold change, y refresca todas las vistas.
        """
        if len(self._manual_clicks) < 4 or self._analizador is None:
            return

        tl = self._manual_clicks[0] # A_sup
        tr = self._manual_clicks[1] # E_sup
        bl = self._manual_clicks[2] # A_inf
        pbs_bot = self._manual_clicks[3] # D_inf (PBS)

        try:
            nuevas = self._analizador.recalibrar_strip(strip_idx, tl, tr, bl, pbs_bot)
        except Exception as e:
            self._actualizar_status_manual(
                strip_idx, f"Error interno: {e}", color="#ff5252")
            self._manual_strip_idx = None
            self._manual_clicks = []
            self._click_markers = []
            return

        if nuevas is None:
            self._actualizar_status_manual(
                strip_idx, "Error al recalibrar.", color="#ff5252")
            self._manual_strip_idx = None
            self._manual_clicks = []
            self._click_markers = []
            return

        # Actualizar la cuantificación de esta tirilla
        self._cuant_lista[strip_idx] = self._analizador.cuantificar(nuevas)

        # Recalcular promedios y fold change
        if len(self._cuant_lista) >= 4:
            try:
                cuant_ctrl = self._cuant_lista[0:2]
                cuant_trat = self._cuant_lista[2:4]
                self._cuant_ctrl_avg = promediar_cuantificaciones(cuant_ctrl)
                self._cuant_trat_avg = promediar_cuantificaciones(cuant_trat)
                self._reporte = calcular_fold_change_con_error(
                    self._cuant_ctrl_avg, self._cuant_trat_avg)
            except Exception as e:
                self._actualizar_status_manual(
                    strip_idx, f"Error: {e}", color="#ff5252")
                return

        # Resetear estado del modo manual ANTES de refrescar vistas
        # (para que la nueva visualización no esté en modo manual)
        self._manual_strip_idx = None
        self._manual_clicks = []
        self._click_markers = []

        # Refrescar todas las vistas
        self._mostrar_reporte()
        self._mostrar_interpretacion()
        self._mostrar_graficos()
        self._mostrar_mapa_visual()  # esto recrea los canvases

        # Status de éxito (después de recrear el label)
        self._actualizar_status_manual(
            strip_idx,
            f"✓ Tirilla {strip_idx + 1} recalibrada manualmente",
            color="#4caf50",
        )

    # ------------------------------------------------------------------ p-valor aprox
    def _p_value_aprox(self, d):
        """
        Test t de Welch aproximado con 2 réplicas internas.
        Es una estimación: con solo 2 réplicas el poder estadístico es bajo,
        pero indica si el cambio es consistente entre las dos tirillas internas.
        """
        try:
            from math import sqrt, isfinite
            nc = d.get('n_ctrl', 1)
            nt = d.get('n_trat', 1)
            if nc < 2 or nt < 2:
                return None
            mc = d['ctrl_norm']
            mt = d['trat_norm']
            sc = d.get('ctrl_std', 0.0)
            st = d.get('trat_std', 0.0)
            # Welch t
            denom = sqrt((sc ** 2) / nc + (st ** 2) / nt)
            if denom < 1e-9:
                return None
            t_stat = abs(mt - mc) / denom
            # Aproximación para 2 vs 2: df ≈ 2; usar tabla simple
            # (Esto es una estimación cruda, con 2 réplicas siempre es poco potente)
            # p ≈ 2 * (1 - cdf(|t|, df=2))
            try:
                from scipy.stats import t as student_t
                df = 2
                p = 2 * (1 - student_t.cdf(t_stat, df))
            except ImportError:
                # Aproximación sin scipy
                p = max(0.001, 1 / (1 + t_stat ** 2))
            if not isfinite(p):
                return None
            return float(p)
        except Exception:
            return None

    # ------------------------------------------------------------------ Excel
    def _exportar(self):
        if not self._reporte or not self._cuant_lista:
            messagebox.showwarning("Sin datos", "Ejecute el análisis primero.")
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            messagebox.showerror(
                "Dependencia faltante", "Instale openpyxl:\n  pip install openpyxl")
            return

        ruta = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="resultados_apoptosis.xlsx")
        if not ruta:
            return

        wb = openpyxl.Workbook()

        # ── Hoja 1: Fold Change ──────────────────────────────────────
        ws = wb.active
        ws.title = "Fold Change"

        hdr_fill = PatternFill("solid", fgColor="1F4E79")
        hdr_font = Font(bold=True, color="FFFFFF")
        thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                      top=Side(style='thin'), bottom=Side(style='thin'))
        centro = Alignment(horizontal="center", vertical="center")

        headers = ["Categoría", "Proteína",
                   "Control (prom. 1-2)", "Control ± desv.",
                   "Tratamiento (prom. 3-4)", "Tratamiento ± desv.",
                   "Veces de cambio", "± error", "p (aprox.)", "¿Qué pasó?"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(1, col, h)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = centro
            c.border = thin

        colores = {'pro-apoptótica': 'FFE0E0',
                   'anti-apoptótica': 'E0E8FF',
                   'otra': 'F2F2F2'}
        # Mapeo a etiquetas legibles para Excel
        etiqueta_tipo = {
            'pro-apoptótica':  'Activadora de muerte celular',
            'anti-apoptótica': 'Bloqueadora de muerte celular',
            'otra':            'Otra',
        }

        row = 2
        for prot in PROTEINAS_UNICAS:
            if prot not in self._reporte:
                continue
            d = self._reporte[prot]
            fill = PatternFill("solid", fgColor=colores.get(d['tipo'], 'FFFFFF'))
            fold_val = None if d['fold_change'] == float('inf') else d['fold_change']
            err_val = d.get('fold_err', None) if fold_val is not None else None
            p_val = self._p_value_aprox(d)
            valores = [etiqueta_tipo.get(d['tipo'], d['tipo']), prot,
                       d['ctrl_norm'], d.get('ctrl_std', 0),
                       d['trat_norm'], d.get('trat_std', 0),
                       fold_val, err_val, p_val,
                       d['estado']]
            for col, v in enumerate(valores, 1):
                c = ws.cell(row, col, v)
                c.fill = fill
                c.border = thin
                c.alignment = Alignment(
                    horizontal="left" if col <= 2 else "center",
                    vertical="center")
            row += 1

        for col_letter, width in zip("ABCDEFGHIJ", [18, 28, 18, 12, 18, 12, 12, 12, 12, 16]):
            ws.column_dimensions[col_letter].width = width
        ws.row_dimensions[1].height = 20

        # Etiquetas para identificar cada tirilla
        etiquetas_tirilla = [
            "Tirilla 1 (Control rep. 1)",
            "Tirilla 2 (Control rep. 2)",
            "Tirilla 3 (Tratamiento rep. 1)",
            "Tirilla 4 (Tratamiento rep. 2)",
        ]

        # ── Hoja 2: Intensidades brutas por tirilla ───────────────────
        ws2 = wb.create_sheet("Intensidades brutas")
        ws2.cell(1, 1, "Proteína").font = Font(bold=True)
        for i in range(len(self._cuant_lista)):
            etiq = etiquetas_tirilla[i] if i < len(etiquetas_tirilla) else f"Tirilla {i+1}"
            ws2.cell(1, i + 2, etiq).font = Font(bold=True)
        for fila_n, prot in enumerate(PROTEINAS_UNICAS, 2):
            ws2.cell(fila_n, 1, prot)
            for col_n, cuant in enumerate(self._cuant_lista, 2):
                bruta = cuant['proteinas'].get(prot, {}).get('intensidad_bruta', "")
                ws2.cell(fila_n, col_n, bruta)
        ws2.column_dimensions['A'].width = 28
        for i in range(len(self._cuant_lista)):
            ws2.column_dimensions[chr(ord('B') + i)].width = 24

        # ── Hoja 3: Intensidades normalizadas por tirilla ─────────────
        ws3 = wb.create_sheet("Intensidades normalizadas")
        ws3.cell(1, 1, "Proteína").font = Font(bold=True)
        for i in range(len(self._cuant_lista)):
            etiq = etiquetas_tirilla[i] if i < len(etiquetas_tirilla) else f"Tirilla {i+1}"
            ws3.cell(1, i + 2, etiq).font = Font(bold=True)
        for fila_n, prot in enumerate(PROTEINAS_UNICAS, 2):
            ws3.cell(fila_n, 1, prot)
            for col_n, cuant in enumerate(self._cuant_lista, 2):
                norm = cuant['proteinas'].get(prot, {}).get('normalizada', "")
                ws3.cell(fila_n, col_n, norm)
        ws3.column_dimensions['A'].width = 28
        for i in range(len(self._cuant_lista)):
            ws3.column_dimensions[chr(ord('B') + i)].width = 24

        wb.save(ruta)
        messagebox.showinfo("Exportado", f"Resultados guardados en:\n{ruta}")

    # ------------------------------------------------------------------ util
    def _log(self, msg: str):
        self.after(0, lambda m=msg: (
            self.textbox.insert(ctk.END, m),
            self.textbox.see(ctk.END),
        ))


if __name__ == "__main__":
    app = App()
    app.mainloop()
